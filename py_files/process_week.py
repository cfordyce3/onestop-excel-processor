# import required libraries
import os
from openpyxl import load_workbook

FILE_DIR = os.path.abspath('..')

# set variables
sheet_day = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# define functions
def row_thru (os,inws,outws) :
    sc = 2
    if   (os == 0):
        sr = 3
    elif (os == 1):
        sr = 12
    elif (os == 2):
        sr = 21
    elif (os == 3):
        sr = 30
    elif (os == 4):
        sr = 39

    r = sr
    c = sc
    for row in inws['B82':'N85']:
        for cell in row:
            outws.cell(row = r, column = c, value = cell.value)
            c += 1

        r += 1
        c = sc

def process_week (week='',month='',year='',show_info=False):

    if (week=='' and month=='' and year==''):
        month=input('What month? '); week=str(input('What week? ')); year=input('What year? ')

    if (show_info == True):
        print()

    out_file = FILE_DIR + '\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb = load_workbook(filename = out_file)

    if (show_info == True):
        print('[' + week + ' ' + month + ', ' + year + ']')

    for os_num in range(0,5):
        in_file = FILE_DIR + '\\Statistics\\'
        if (os_num == 0):
            if (show_info == True):
                print('Processing Low Desk... ', end='')
            in_file += 'Low Desk\\' + year + '\\' + month + '\\LD Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
        else:
            if (show_info == True):
                print('Processing OS' + str(os_num) + '... ', end='')
            in_file += 'OS' + str(os_num) + '\\' + year + '\\' + month + '\\OS' + str(os_num) + ' Week of ' + month + ' ' + week + ', ' + year + '.xlsx'

        try:
            #print(in_file)
            f = open(in_file)
            f.close()
        except FileNotFoundError:
            if (show_info == True):
                print('FILE NOT FOUND. DATA NOT COMPILED!')
            continue

        inwb = load_workbook(filename = in_file, data_only = True)
        for sheet_num in sheet_day:
            inws = inwb[sheet_num]
            outws = outwb[sheet_num]
            row_thru(os_num,inws,outws)

        if (show_info == True):
            print('Done.')

    # determine save file and saves file
    save_file = FILE_DIR + '\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb.save(save_file)

#process_week('26','August','2019')
