from openpyxl import load_workbook

sheet_day = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

def row_thru_error (os, ws, show_info=False):
    # define starting and finishing rows and columns
    sc = 2; fc = 14
    if (os == 0):
        sr = 3; fr = 6
    elif (os == 1):
        sr = 12; fr = 15
    elif (os == 2):
        sr = 21; fr = 24
    elif (os == 3):
        sr = 30; fr = 33
    elif (os == 4):
        sr = 39; fr = 42

    error_count = 0
    for row in ws.iter_rows(min_row=sr, min_col=sc, max_col=fc, max_row=fr):
        for cell in row:
            if (type(cell.value) != int):
                error_count+=1
                if (show_info==True):
                    print('error: ' + str(cell.value) + ' at ' + str(cell.coordinate))
    return (error_count > 0)

def row_thru_correct (os,inws,outws) :
    sc = 2
    if   (os == 0):
        sr = 3
    elif (os == 1):
        sr = 12
    elif (os == 2):
        sr = 21
    elif (os == 3):
        sr = 30
    elif (os == 4):
        sr = 39

    # find correct starting place
    check1 = False; check2 = False; col_coord=0; row_coord=0
    for row in inws.iter_rows(min_row=50, min_col=2, max_col=2, max_row=150):
        if (row_coord != 0):
            break
        for cell in row:
            if (check1 == True and check2 == True):
                col_coord=cell.column
                row_coord=cell.row
                break
            if (cell.value == 'Bursar'):
                check1 = True
            if (cell.value == 'Walk-Ins' and check1 == True):
                check2 = True

    r = sr
    c = sc
    for row in inws.iter_rows(min_row=row_coord, min_col=col_coord, max_col=col_coord+12, max_row=row_coord+3):
        for cell in row:
            outws.cell(row = r, column = c, value = cell.value)
            c += 1
        r += 1
        c = sc

def error_correct (week,month,year):

    out_file = 'C:\\Users\\fordy\\Desktop\\One Stop GitHub\\onestop-excel-processor\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb = load_workbook(filename = out_file)

    for os_num in range(0,5):
        in_file = 'C:\\Users\\fordy\\Desktop\\One Stop GitHub\\onestop-excel-processor\\Statistics\\'
        if (os_num == 0):
            in_file += 'Low Desk\\' + year + '\\' + month + '\\LD Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
        else:
            in_file += 'OS' + str(os_num) + '\\' + year + '\\' + month + '\\OS' + str(os_num) + ' Week of ' + month + ' ' + week + ', ' + year + '.xlsx'

        inwb = load_workbook(filename = in_file, data_only = True)
        for sheet_num in sheet_day:
            inws = inwb[sheet_num]
            outws = outwb[sheet_num]
            row_thru_correct(os_num,inws,outws)

    save_file = 'C:\\Users\\fordy\\Desktop\\One Stop GitHub\\onestop-excel-processor\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb.save(save_file)

def error_check (day='', month='', year='',show_info=False):

    error_found = False

    file = 'C:\\Users\\fordy\\Desktop\\One Stop GitHub\\onestop-excel-processor\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\Week of ' + month + ' ' + day + ', ' + year + '.xlsx'
    wb = load_workbook(filename=file, data_only=True)

    for os_num in range(0,5):
        for sheet in sheet_day:
            ws = wb[sheet]
            if (show_info==True):
                print(str(sheet + ' OS num: ' + str(os_num)), end=' ')
            if (row_thru_error(os_num,ws) == True):
                if (show_info==True):
                    print('Error found.',end='')
                error_found = True
            if (show_info==True):
                print()

    return error_found

#wb = load_workbook(filename='C:\\Users\\fordy\\Desktop\\One Stop GitHub\\onestop-excel-processor\\Statistics\\Master\\2019\\Weekly\\August\\Week of August 12, 2019.xlsx', data_only=True)
#ws = wb['Wednesday']
#row_thru_correct(0,ws,True)

#error_check('19','August','2019',True)

error_correct('19','August','2019')
