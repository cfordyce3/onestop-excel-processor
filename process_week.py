# import required libraries
import os
from openpyxl import load_workbook
from resources import get_stats_folder

# set working directory to the directory above where this file is located
FILE_DIR = get_stats_folder()

# set variables
sheet_day = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# define functions
def row_thru (os,inws,outws) :
    sc = 2
    if   (os == 0):
        sr = 3
    elif (os == 1):
        sr = 12
    elif (os == 2):
        sr = 21
    elif (os == 3):
        sr = 30
    elif (os == 4):
        sr = 39

    # find correct starting place
    check1 = False; check2 = False; col_coord=0; row_coord=0
    for row in inws.iter_rows(min_row=50, min_col=2, max_col=2, max_row=150):
        if (row_coord != 0):
            break
        for cell in row:
            if (check1 == True and check2 == True):
                col_coord=cell.column
                row_coord=cell.row
                break
            if (cell.value == 'Bursar'):
                check1 = True
            if (cell.value == 'Walk-Ins' and check1 == True):
                check2 = True

    r = sr
    c = sc
    for row in inws.iter_rows(min_row=row_coord, min_col=col_coord, max_col=col_coord+12, max_row=row_coord+3):
        for cell in row:
            outws.cell(row = r, column = c, value = cell.value)
            c += 1
        r += 1
        c = sc


# correct the totals at the end
def correct_totals (wb):
    # correct the daily sheets
    for sheet_name in sheet_day:
        ws = wb[sheet_name]
        for x in range(2,15):
            for y in range(51,55):
                ws.cell(row=y,column=x).value = ws.cell(row=y-48,column=x).value + ws.cell(row=y-39,column=x).value + ws.cell(row=y-30,column=x).value + ws.cell(row=y-21,column=x).value + ws.cell(row=y-12,column=x).value

    # correct the weekly stats section up to the final total section
    multiplier = 0
    for sheet_name in sheet_day:
        ws = wb['Weekly Stats']

        for x in range(2,15):
            y_source = 51
            for y in range(3+(multiplier*7),7+(multiplier*7)):
                if (y != (7+(multiplier*7)-1)):
                    ws.cell(row=y,column=x).value = wb[sheet_name].cell(row=y_source,column=x).value
                    y_source += 1
                else:
                    ws.cell(row=y,column=x).value = ws.cell(row=y-3,column=x).value + ws.cell(row=y-2,column=x).value + ws.cell(row=y-1,column=x).value
        multiplier += 1

    # set up final total section at the bottom of the sheet
    ws = wb['Weekly Stats']
    for x in range(2,15):
        for y in range(38,42):
            if (y != 41):
                ws.cell(row=y,column=x).value = ws.cell(row=y-35,column=x).value + ws.cell(row=y-28,column=x).value + ws.cell(row=y-21,column=x).value + ws.cell(row=y-14,column=x).value + ws.cell(row=y-7,column=x).value
            else:
                ws.cell(row=y,column=x).value = ws.cell(row=y-3,column=x).value + ws.cell(row=y-2,column=x).value + ws.cell(row=y-1,column=x).value


def process_week (week='',month='',year='',show_info=False):
    if (show_info == True):
        print()

    out_file = FILE_DIR + '\\Statistics\\Master\\' + '\\Weekly\\' + year + '\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb = load_workbook(filename = out_file)

    if (show_info == True):
        print('[' + week + ' ' + month + ', ' + year + ']')

    for os_num in range(0,5):
        in_file = FILE_DIR + '\\Statistics\\'
        if (os_num == 0):
            if (show_info == True):
                print('Processing Low Desk... ', end='')
            in_file += 'Low Desk\\' + year + '\\' + month + '\\LD Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
        else:
            if (show_info == True):
                print('Processing OS' + str(os_num) + '... ', end='')
            in_file += 'OS' + str(os_num) + '\\' + year + '\\' + month + '\\OS' + str(os_num) + ' Week of ' + month + ' ' + week + ', ' + year + '.xlsx'

        try:
            f = open(in_file)
            f.close()
        except FileNotFoundError:
            if (show_info == True):
                print('FILE NOT FOUND. DATA NOT COMPILED!')
            continue

        inwb = load_workbook(filename = in_file, data_only = True)
        for sheet_name in sheet_day:
            inws = inwb[sheet_name]
            outws = outwb[sheet_name]
            row_thru(os_num,inws,outws)

        if (show_info == True):
            print('Done.')


    correct_totals(outwb)

    # determine save file and saves file
    save_file = FILE_DIR + '\\Statistics\\Master\\' + '\\Weekly\\' + year + '\\' + month + '\\Week of ' + month + ' ' + week + ', ' + year + '.xlsx'
    outwb.save(save_file)

#process_week('5','August','2019')
