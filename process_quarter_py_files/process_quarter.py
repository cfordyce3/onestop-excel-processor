import os
from openpyxl import Workbook, load_workbook
from create_spreadsheet import create_spreadsheet

FILE_DIR = os.path.abspath('..')
months = {'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,'July':7,'August':8,'September':9,'October':10,'November':11,'December':12}

def strip_to_day (filename,month_name,year_name):
    day = filename.strip('Week of ' + month_name + ' ')
    day = day.rstrip(str(year_name) + '.xlsx')
    day = day.strip(", ")
    return day

def sort_weeks (weeks,month_name,year_name):
    sorted_weeks = []

    for week in range(0,len(weeks)):
        weeks[week] = int(strip_to_day(weeks[week],month_name,year_name))

    weeks.sort()

    for week in range(0,len(weeks)):
        day = str(weeks[week])
        sorted_weeks.append('Week of ' + month_name + ' ' + day + ', ' + year_name + '.xlsx')

    return sorted_weeks

def create_dict_of_inputs (start_week='',start_month='',start_year='',end_week='',end_month='',end_year='',show_info=False):
    start_month_val = months[start_month]
    end_month_val = months[end_month]

    inputs = {}

    if (start_year==end_year):
        year_list = {} # dictionary of years (e.g. {2019:..., 2020:...})
        month_list = {} # dictionary of months (e.g. {'January':...,'February':...}
        for month_name,month_num in months.items(): # go through each month
            day_list = [] # list of specific week per month (e.g. ['Week of January 1, 2019.xlsx', 'Week of January 8, 2019.xlsx'...])
            if (month_num >= start_month_val and month_num <= end_month_val): # if the month is within the starting month and the ending month (one year only)
                for file in os.listdir(FILE_DIR + '\\Statistics\\Master\\' + str(start_year) + '\\Weekly\\' + month_name):
                    day = strip_to_day(file,month_name,start_year)
                    if (month_name != start_month and month_name != end_month):
                        day_list.append(file)
                    elif (month_name == end_month and int(day) <= int(end_week)):
                        day_list.append(file)
                    elif (month_name == start_month and int(day) >= int(start_week)):
                        day_list.append(file)
                if (len(day_list) != 0):
                    month_list[month_name] = sort_weeks(day_list,month_name,start_year)
            if (len(month_list) != 0):
                year_list[start_year] = month_list
        inputs = year_list

    else: # start_year != end_year
        year_list = {} # dictionary of years (e.g. {2019:..., 2020:...})
        month_list = {} # dictionary of months (e.g. {'January':...,'February':...}
        for month_name,month_num in months.items(): # go through each month
            day_list = [] # list of specific week per month (e.g. ['Week of January 1, 2019.xlsx', 'Week of January 8, 2019.xlsx'...])
            if (month_num >= start_month_val and month_num <= end_month_val): # if the month is within the starting month and the ending month (one year only)
                for file in os.listdir(FILE_DIR + '\\Statistics\\Master\\' + str(start_year) + '\\Weekly\\' + month_name):
                    day = strip_to_day(file,month_name,start_year)
                    if (month_name != start_month and month_name != end_month):
                        day_list.append(file)
                    elif (month_name == end_month and int(day) <= int(end_week)):
                        day_list.append(file)
                    elif (month_name == start_month and int(day) >= int(start_week)):
                        day_list.append(file)
                if (len(day_list) != 0):
                    month_list[month_name] = sort_weeks(day_list,month_name,start_year)
            if (len(month_list) != 0):
                year_list[start_year] = month_list
        inputs = year_list

    if (show_info == True): print(inputs)

    return inputs

def row_thru (outwb,year,month,filename,first,show_info=False):
    in_file = FILE_DIR + '\\Statistics\\Master\\' + year + '\\Weekly\\' + month + '\\' + filename
    inwb = load_workbook(filename=in_file,data_only=True)
    inws = inwb['Weekly Stats']

    ws_name_from_filename = filename.rstrip('.xlsx')
    outws = None
    if (first==True): outws = outwb.active; outws.title = ws_name_from_filename
    else: outws = outwb.create_sheet(title=ws_name_from_filename)

    if (show_info==True): print(outws.title)
    create_spreadsheet(outws)

    # copy the data
    r = 3; c = 2
    for row in inws['B38':'N41']:
        for cell in row:
            if (show_info==True and row==3): print('Input: ')
            if (show_info==True): print(cell.coordinate + ' ' + str(cell.value),end=' ')
            outws.cell(row=r,column=c,value=cell.value)
            c+=1
        r+=1
        c=2
        if (show_info==True): print()
    if (show_info==True): input('Continue? ')

def process_quarter (season='',start_week='',start_month='',start_year='',end_week='',end_month='',end_year='',show_info=False):
    if (season=='' and start_week=='' and start_month=='' and start_year=='' and end_week=='' and end_month=='' and end_year==''):
        start_month=input('What starting month? '); start_week=str(input('What starting week? ')); start_year=input('What starting year? ')
        end_month=input('What ending month? '); end_week=str(input('What ending week? ')); end_year=input('What ending year? ')
        season=input('What season are we processing? ')

    # set the output file directory and load the workbook
    out_file_name = FILE_DIR + '\\Statistics\\Master\\' + end_year + '\\Quarterly\\' + season + ' ' + end_year + '.xlsx'
    outwb = Workbook()

    # retrieve the files between the dates selected
    input_files = create_dict_of_inputs(start_week,start_month,start_year,end_week,end_month,end_year)
    first = True
    for year,month_list in input_files.items():
        for month,actual_files in month_list.items():
            for file in actual_files:
                if (show_info==True): print('Processing: ' + file,end=' ... '); print()
                row_thru(outwb,year,month,file,first,show_info)
                if (show_info==True): print('Done.')
                first=False

    outwb.save(out_file_name)


#create_dict_of_inputs('20','June','2019','9','September','2019',True)
process_quarter('Summer_test','20','June','2019','9','September','2019',False)
#process_quarter('Test','5','August','2019','26','August','2019',False)
