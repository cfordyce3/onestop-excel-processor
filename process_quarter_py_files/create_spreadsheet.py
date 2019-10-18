from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment, NamedStyle
from openpyxl.chart import PieChart3D, Reference
from openpyxl.chart.label import DataLabelList

def set_default_variables (ws):
    # vertical
    ws['A1'] = 'Total'
    ws['A3'] = 'Students'
    ws['A4'] = 'Parents'
    ws['A5'] = 'Other'
    ws['A6'] = 'Total'

    ws.merge_cells('A9:B9')
    ws['A9'] = 'Total'
    ws['A10'] = 'Bursar'
    ws['A11'] = 'Financial Aid'
    ws['A12'] = 'Registrar'
    ws['A13'] = 'Other'

    ws.merge_cells('A15:B15')
    ws['A15'] = 'Walk-Ins'
    ws['A16'] = 'Bursar'
    ws['A17'] = 'Financial Aid'
    ws['A18'] = 'Registrar'
    ws['A19'] = 'Other'

    ws.merge_cells('A21:B21')
    ws['A21'] = 'Phones'
    ws['A22'] = 'Bursar'
    ws['A23'] = 'Financial Aid'
    ws['A24'] = 'Registrar'
    ws['A25'] = 'Other'

    ws.merge_cells('A27:B27')
    ws['A27'] = 'Emails'
    ws['A28'] = 'Bursar'
    ws['A29'] = 'Financial Aid'
    ws['A30'] = 'Registrar'
    ws['A31'] = 'Other'

    # horizontal | first row
    ws.merge_cells('B1:D1')
    ws.merge_cells('E1:G1')
    ws.merge_cells('H1:J1')
    ws.merge_cells('K1:M1')
    ws['B1'] = 'Bursar'
    ws['E1'] = 'Financial Aid'
    ws['H1'] = 'Registrar'
    ws['K1'] = 'Other'

    # horizontal | second row
    ws['B2'] = 'Walk-Ins'
    ws['C2'] = 'Phones'
    ws['D2'] = 'Emails'
    ws['E2'] = 'Walk-Ins'
    ws['F2'] = 'Phones'
    ws['G2'] = 'Emails'
    ws['H2'] = 'Walk-Ins'
    ws['I2'] = 'Phones'
    ws['J2'] = 'Emails'
    ws['K2'] = 'Walk-Ins'
    ws['L2'] = 'Phones'
    ws['M2'] = 'Emails'

def set_default_colors (ws):
    ws['A1'].font = Font(size=14,bold=True,underline='single')
    ws['A1'].fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('B1A0C7'))

    first_row = ['B1','E1','H1','K1']
    for cell in first_row:
        ws[cell].alignment = Alignment(horizontal='center',vertical='bottom')
        ws[cell].fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('92CDDC'))


    second_row_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('C4D79B'))
    for col in range(2,14):
        ws.cell(column=col,row=2).fill = second_row_fill

    first_col_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('FABF8F'))
    for row in range(3,7):
        ws.cell(column=1,row=row).fill = first_col_fill

    white_font = Font(color='FFFFFF')

    blacks = [9,15,21,27]
    for row in blacks:
        ws.cell(column=1,row=row).alignment = Alignment(horizontal='center',vertical='bottom')
        ws.cell(column=1,row=row).font = white_font
        ws.cell(column=1,row=row).fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('000000'))

    blue_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('538DD5'))
    blues = [10,16,22,28]
    for row in blues:
        ws.cell(column=1,row=row).fill = blue_fill
        ws.cell(column=1,row=row).font = white_font

    red_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('C00000'))
    reds = [11,17,23,29]
    for row in reds:
        ws.cell(column=1,row=row).fill = red_fill
        ws.cell(column=1,row=row).font = white_font

    green_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('76933C'))
    greens = [12,18,24,30]
    for row in greens:
        ws.cell(column=1,row=row).fill = green_fill
        ws.cell(column=1,row=row).font = white_font

    purple_fill = PatternFill(patternType='solid',fill_type='solid',fgColor=Color('6F30A0'))
    purples = [13,19,25,31]
    for row in purples:
        ws.cell(column=1,row=row).fill = purple_fill
        ws.cell(column=1,row=row).font = white_font

def set_default_equations (ws):
    # totals
    ws['B10'] = '=SUM(B6:D6)'
    ws['B11'] = '=SUM(E6:G6)'
    ws['B12'] = '=SUM(H6:J6)'
    ws['B13'] = '=SUM(K6:M6)'

    # walk-ins
    ws['B16'] = '=B6'
    ws['B17'] = '=E6'
    ws['B18'] = '=H6'
    ws['B19'] = '=K6'

    # phones
    ws['B22'] = '=C6'
    ws['B23'] = '=F6'
    ws['B24'] = '=I6'
    ws['B25'] = '=L6'

    # emails
    ws['B28'] = '=D6'
    ws['B29'] = '=G6'
    ws['B30'] = '=J6'
    ws['B31'] = '=M6'

def create_pie_chart (ws):
    pie_chart = PieChart3D()

    labels = Reference(ws,min_col=1,min_row=10,max_row=13)
    data = Reference(ws,min_col=2,min_row=10,max_row=13)

    pie_chart.add_data(data,titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.legend.position = 'b'
    pie_chart.title = ''
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True

    ws.add_chart(pie_chart,'D9')

def create_spreadsheet (ws):
    # set values for first table
    set_default_variables(ws)

    # set colors for first table
    set_default_colors(ws)

    # set the equations for the spreadsheet to use later
    set_default_equations(ws)

    # create the pie chart
    create_pie_chart(ws)
